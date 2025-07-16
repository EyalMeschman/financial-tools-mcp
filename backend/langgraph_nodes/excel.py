"""Excel node for report generation."""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.azure_adapter import InvoiceData


def invoice_suffix(invoice: InvoiceData) -> str:
    """Extract invoice suffix from filename per spec.

    - Strip non-digits, take last 4, left-pad zeros
    - If no digits → SUFFIX_NOT_FOUND

    Args:
        invoice: Invoice data (unused in current implementation)
        filename: Original filename

    Returns:
        str: Invoice suffix (4 digits or SUFFIX_NOT_FOUND)
    """
    if not invoice.InvoiceId or not invoice.InvoiceId.content:
        return "SUFFIX_NOT_FOUND"

    # Strip non-digits and extract all digits
    digits = re.sub(r"\D", "", invoice.InvoiceId.content)

    if not digits:
        return "SUFFIX_NOT_FOUND"

    # Take last 4 digits, left-pad with zeros if needed
    last_four = digits[-4:]
    return last_four.zfill(4)


async def run(input: dict) -> dict:
    """Excel node that generates Excel reports from invoice data.

    Args:
        input: Input dictionary containing 'invoices' key with list of InvoiceData objects
              and 'target_currency' key

    Returns:
        dict: Dictionary with 'xlsx' (BytesIO bytes) and 'row_count' keys
    """
    invoices: list[InvoiceData] = input.get("invoices", [])
    target_currency: str = input.get("target_currency", "USD")

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices Report"

    # Define headers per spec
    headers = [
        "Date (DD/MM/YYYY)",
        "Invoice Suffix",
        f"{target_currency} Total Price",
        "Foreign Currency Total Price",
        "Foreign Currency Code",
        "Exchange Rate",
        "Vendor Name",
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    row_num = 2
    filename = ""  # Initialize filename for empty invoices case
    for invoice in invoices:
        # Extract filename from invoice data (should be stored in invoice)
        filename = getattr(invoice, "_filename", "")

        # Extract data with error handling per spec
        date = invoice.InvoiceDate.value_date.strftime("%d-%m-%Y") if invoice.InvoiceDate else "ERROR"
        suffix = invoice_suffix(invoice)
        vendor_name = invoice.VendorName.content if invoice.VendorName else "N/A"

        # Handle amounts and currency
        target_total = "N/A"
        foreign_total = ""
        foreign_currency = ""
        exchange_rate = ""

        if invoice.InvoiceTotal and invoice.InvoiceTotal.value_currency:
            invoice_currency = invoice.InvoiceTotal.value_currency.currency_code
            invoice_amount = invoice.InvoiceTotal.value_currency.amount

            if invoice_currency == target_currency:
                # Case 1: Same currency - use invoice amount directly, leave foreign fields empty
                target_total = invoice_amount
                foreign_total = ""
                foreign_currency = ""
                exchange_rate = ""
            else:
                # Case 2: Different currency - fill all foreign fields
                foreign_total = invoice_amount
                foreign_currency = invoice_currency

                # Use converted amounts if available
                target_total = getattr(invoice, "_converted_amount", "N/A")
                exchange_rate = getattr(invoice, "_exchange_rate", "N/A")

                # Format exchange rate to 4 decimal places if it's a number
                if isinstance(exchange_rate, int | float):
                    exchange_rate = f"{exchange_rate:.4f}"

        # Write row data
        row_data = [date, suffix, target_total, foreign_total, foreign_currency, exchange_rate, vendor_name]

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)

        row_num += 1

    # Add placeholder ERROR rows if no invoices provided
    if not invoices:
        error_row = ["ERROR", filename, "N/A", "N/A", "N/A", "N/A", "N/A"]
        for col, value in enumerate(error_row, 1):
            ws.cell(row=2, column=col, value=value)
        row_num = 3

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Also save to exports directory if job_id is provided
    job_id = input.get("job_id")
    if job_id:
        from pathlib import Path

        exports_dir = Path("exports")
        exports_dir.mkdir(exist_ok=True)

        export_path = exports_dir / f"{job_id}.xlsx"
        wb.save(export_path)

    # Calculate row count (excluding header)
    data_row_count = max(1, len(invoices)) if invoices else 1

    return {"xlsx": excel_buffer.getvalue(), "row_count": data_row_count}

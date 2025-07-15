"""Tests for the excel node."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.azure_adapter import DefaultContent, InvoiceData, InvoiceTotal, ValueCurrency
from langgraph_nodes import excel


class TestExcelNode:
    """Test cases for the excel node."""

    def test_module_imports(self):
        """Test that the excel module imports successfully."""
        assert excel is not None

    def test_invoice_suffix_helper(self):
        """Test the invoice_suffix helper function per spec."""
        # Test with digits - take last 4, left-pad zeros
        mock_invoice = InvoiceData(
            InvoiceId=DefaultContent(content="INV-123", confidence=0.95),
            InvoiceDate=None,
            VendorName=None,
            VendorAddressRecipient=None,
            InvoiceTotal=None,
        )
        assert excel.invoice_suffix(mock_invoice) == "0123"

        mock_invoice.InvoiceId.content = "456789"
        assert excel.invoice_suffix(mock_invoice) == "6789"

        mock_invoice.InvoiceId.content = "INV-2024-001"
        assert excel.invoice_suffix(mock_invoice) == "4001"

        # Test with no digits
        mock_invoice.InvoiceId.content = "INV-NO-DIGITS"
        assert excel.invoice_suffix(mock_invoice) == "SUFFIX_NOT_FOUND"

        # Test empty/None InvoiceId
        mock_invoice.InvoiceId.content = ""
        assert excel.invoice_suffix(mock_invoice) == "SUFFIX_NOT_FOUND"

        mock_invoice.InvoiceId = None
        assert excel.invoice_suffix(mock_invoice) == "SUFFIX_NOT_FOUND"

    @pytest.mark.asyncio
    async def test_run_with_empty_invoices(self):
        """Test run function with empty invoice list."""
        result = await excel.run({"target_currency": "USD"})

        # Check return structure
        assert isinstance(result, dict)
        assert "xlsx" in result
        assert "row_count" in result
        assert isinstance(result["xlsx"], bytes)
        assert result["row_count"] == 1  # One ERROR row

        # Validate Excel content
        excel_data = BytesIO(result["xlsx"])
        wb = load_workbook(excel_data)
        ws = wb.active

        # Check headers per spec
        expected_headers = [
            "Date (DD/MM/YYYY)",
            "Invoice Suffix",
            "USD Total Price",
            "Foreign Currency Total Price",
            "Foreign Currency Code",
            "Exchange Rate (4 dp)",
            "Vendor Name",
        ]
        for col, expected_header in enumerate(expected_headers, 1):
            assert ws.cell(row=1, column=col).value == expected_header

        # Check ERROR row per spec
        expected_error_row = ["ERROR", "No invoices found", "N/A", "N/A", "N/A", "N/A", "N/A"]
        for col, expected_value in enumerate(expected_error_row, 1):
            assert ws.cell(row=2, column=col).value == expected_value

    @pytest.mark.asyncio
    async def test_run_with_complete_invoice(self):
        """Test run function with a complete invoice."""
        # Create test invoice data
        invoice = InvoiceData(
            InvoiceId=DefaultContent(content="INV-001", confidence=0.95),
            InvoiceDate=DefaultContent(content="2024-01-15", confidence=0.95),
            VendorName=DefaultContent(content="Acme Corp", confidence=0.95),
            VendorAddressRecipient=DefaultContent(content="123 Business St", confidence=0.95),
            InvoiceTotal=InvoiceTotal(
                value_currency=ValueCurrency(amount=1234.56, currency_code="EUR"), content="1,234.56", confidence=0.95
            ),
        )

        # Add filename attribute for testing
        invoice._filename = "test_invoice_001.pdf"

        result = await excel.run({"invoices": [invoice], "target_currency": "USD"})

        # Check return structure
        assert isinstance(result, dict)
        assert "xlsx" in result
        assert "row_count" in result
        assert isinstance(result["xlsx"], bytes)
        assert result["row_count"] == 1

        # Validate Excel content
        excel_data = BytesIO(result["xlsx"])
        wb = load_workbook(excel_data)
        ws = wb.active

        # Check data row per spec (different currency: EUR vs USD)
        assert ws.cell(row=2, column=1).value == "2024-01-15"  # Date
        assert ws.cell(row=2, column=2).value == "0001"  # Invoice Suffix (from InvoiceId)
        assert ws.cell(row=2, column=3).value == "N/A"  # USD Total Price (not converted yet)
        assert ws.cell(row=2, column=4).value == 1234.56  # Foreign Currency Total Price
        assert ws.cell(row=2, column=5).value == "EUR"  # Foreign Currency Code
        assert ws.cell(row=2, column=6).value == "N/A"  # Exchange Rate (not converted yet)
        assert ws.cell(row=2, column=7).value == "Acme Corp"  # Vendor Name

    @pytest.mark.asyncio
    async def test_run_with_partial_invoice(self):
        """Test run function with partial invoice data (missing fields)."""
        # Create invoice with missing fields
        invoice = InvoiceData(
            InvoiceId=DefaultContent(content="INV-002", confidence=0.95),
            InvoiceDate=None,  # Missing date
            VendorName=DefaultContent(content="Test Vendor", confidence=0.95),
            VendorAddressRecipient=None,  # Missing address
            InvoiceTotal=None,  # Missing total
        )

        # Add filename with no digits
        invoice._filename = "test_invoice.pdf"

        result = await excel.run({"invoices": [invoice], "target_currency": "GBP"})

        # Validate Excel content
        excel_data = BytesIO(result["xlsx"])
        wb = load_workbook(excel_data)
        ws = wb.active

        # Check data row with ERROR/N/A placeholders per spec
        assert ws.cell(row=2, column=1).value == "ERROR"  # Date (missing)
        assert ws.cell(row=2, column=2).value == "0002"  # Invoice Suffix (from InvoiceId)
        assert ws.cell(row=2, column=3).value == "N/A"  # GBP Total Price (missing)
        assert ws.cell(row=2, column=4).value in ["", None]  # Foreign Currency Total Price (missing)
        assert ws.cell(row=2, column=5).value in ["", None]  # Foreign Currency Code (missing)
        assert ws.cell(row=2, column=6).value in ["", None]  # Exchange Rate (missing)
        assert ws.cell(row=2, column=7).value == "Test Vendor"  # Vendor Name

    @pytest.mark.asyncio
    async def test_run_with_multiple_invoices(self):
        """Test run function with multiple invoices."""
        # Create multiple test invoices
        invoice1 = InvoiceData(
            InvoiceId=DefaultContent(content="INV-001", confidence=0.95),
            InvoiceDate=DefaultContent(content="2024-01-15", confidence=0.95),
            VendorName=DefaultContent(content="Vendor A", confidence=0.95),
            VendorAddressRecipient=DefaultContent(content="Address A", confidence=0.95),
            InvoiceTotal=InvoiceTotal(
                value_currency=ValueCurrency(amount=100.0, currency_code="USD"), content="100.00", confidence=0.95
            ),
        )
        invoice1._filename = "invoice_123.pdf"

        invoice2 = InvoiceData(
            InvoiceId=DefaultContent(content="INV-002", confidence=0.95),
            InvoiceDate=DefaultContent(content="2024-01-16", confidence=0.95),
            VendorName=DefaultContent(content="Vendor B", confidence=0.95),
            VendorAddressRecipient=DefaultContent(content="Address B", confidence=0.95),
            InvoiceTotal=InvoiceTotal(
                value_currency=ValueCurrency(amount=200.0, currency_code="EUR"), content="200.00", confidence=0.95
            ),
        )
        invoice2._filename = "invoice_456.pdf"

        result = await excel.run({"invoices": [invoice1, invoice2], "target_currency": "CAD"})

        # Check return structure
        assert result["row_count"] == 2

        # Validate Excel content
        excel_data = BytesIO(result["xlsx"])
        wb = load_workbook(excel_data)
        ws = wb.active

        # Check first invoice
        assert ws.cell(row=2, column=1).value == "2024-01-15"  # Date
        assert ws.cell(row=2, column=2).value == "0001"  # Invoice Suffix
        assert ws.cell(row=2, column=5).value == "USD"  # Foreign Currency Code
        assert ws.cell(row=2, column=7).value == "Vendor A"  # Vendor Name

        # Check second invoice
        assert ws.cell(row=3, column=1).value == "2024-01-16"  # Date
        assert ws.cell(row=3, column=2).value == "0002"  # Invoice Suffix
        assert ws.cell(row=3, column=5).value == "EUR"  # Foreign Currency Code
        assert ws.cell(row=3, column=7).value == "Vendor B"  # Vendor Name

    @pytest.mark.asyncio
    async def test_excel_formatting(self):
        """Test that Excel file has proper formatting."""
        invoice = InvoiceData(
            InvoiceId=DefaultContent(content="INV-001", confidence=0.95),
            InvoiceDate=DefaultContent(content="2024-01-15", confidence=0.95),
            VendorName=DefaultContent(content="Test Vendor", confidence=0.95),
            VendorAddressRecipient=DefaultContent(content="Test Address", confidence=0.95),
            InvoiceTotal=InvoiceTotal(
                value_currency=ValueCurrency(amount=100.0, currency_code="USD"), content="100.00", confidence=0.95
            ),
        )
        invoice._filename = "test_123.pdf"

        result = await excel.run({"invoices": [invoice], "target_currency": "EUR"})

        # Validate Excel formatting
        excel_data = BytesIO(result["xlsx"])
        wb = load_workbook(excel_data)
        ws = wb.active

        # Check that worksheet has correct title per spec
        assert ws.title == "Invoices Report"

        # Check header formatting (basic checks)
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.value == "Date (DD/MM/YYYY)"
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb == "00FFFFFF"
